###这个是对第一轮筛查完成后再次调整的生成的词，做第二次替换，并且重新生成词表的程序

import csv
import re
import string
from array import array

import numpy
import numpy as np
import openpyxl
import sklearn
import nltk
from nltk import WordNetLemmatizer, MWETokenizer
from nltk.corpus import stopwords
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import CountVectorizer, TfidfTransformer
print('Please input testnum and year-to-year:')
testnum=list(map(int,input().split()))


punctuation_map=dict((ord(char),' ') for char in string.punctuation if not '-')
text=[]
wordnum=[]
s=nltk.stem.SnowballStemmer('english')
stopw=stopwords.words('english')

readbook = openpyxl.load_workbook('D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\可能的词组2.5_2022.04.10更新.xlsx')
sheet=readbook.get_sheet_names()
ws=readbook.get_sheet_by_name(sheet[0])

cizu=[]
for w in range(2,ws.max_row+1):
    prase=ws.cell(w,1).value
    if prase is not None:
        p=prase.split(' ')
        pp=tuple(p)
        cizu.append(pp)

def get_wordnet_pos(treebank_tag):
    if treebank_tag.startswith('J'):
        return nltk.corpus.wordnet.ADJ
    elif treebank_tag.startswith('V'):
        return nltk.corpus.wordnet.VERB
    elif treebank_tag.startswith('N'):
        return nltk.corpus.wordnet.NOUN
    elif treebank_tag.startswith('R'):
        return nltk.corpus.wordnet.ADV
lmtzr=WordNetLemmatizer()


replace=['libraries','metasearching','vigorously','absys','al','il','the','sisissunrise',
'sirsidynixs','dynixs','librarysirsidynix','sirsis','12','apis','wms','instal','installed','librarysolution','vsmart','nextgeneration','lsp','asp','libriss','opal','nonils','oclcs','forprofit','indexbased','newgeneration','exclusively','lsps','taylor','caput','arl','pcbased','ill','librarys','eresource','patronfacing','softwareasaservice','platforms','bowker','cataloging','fsc','folletts','selfservice','cig','fretwelldowning','librarythe','innovatives','ukbased','products','productsthe','servicesthe','signon','vdx','vinsight','eds','sky','webscale','systemthe','shemtov','servicein','serviceoriented','servicethis','searching','selfcheck','selfcheckout','selfhosted','serf','relevancybased','relevancyranked','oers','makerspaces','ncsu','libraryacademic','librarybibliocommons','librarycardcom','librarycomcom','librarycompanion','librarycompany','libraryex','libraryoriented','librarypublic','libraryspecial','lbs','informationssysteme','libraryin','wellestablished','fullfeatured','industrythe','innreach','tao','webnative','cloudbased','librarything','cdi','migrating','nonus','serviceoriented','familyowned','firstsearch','firsttime','geacs','frbr','ftc','gi','libraryprovided','libraryrelated','longerterm','kansa','ncip','nypl','opacs','yearend','amsterdambased','vtlss','eresources','digitalfirst','everincreasing','librarians','ilsthe','integrate','librarydecision','librarylibrary','libraryon','peertopeer','sectorthe','subscriptionbased','soa','toptier','urm','allwebbased','alm','almaex','bookit','broadbased','companyin','developmentin','dosbased','dbtextworks','dras','evergrowing','headquarters','impactonline','impactverso','eosi','hostterminal','migrate'
]
replace_word=['library','metasearch','vigorous','absysNet','alS','ils',' ','sisis sunrise','sirsidynix','dynix','library sirsidynix','sirsidynix','k12','api','worldshare management service','installation','installation','library solution','v smart','next generation','library service platform','application service provider','libris','opals','non ils','oclc','for profit','index based','new generation','exclusive','library service platform','baker taylor','capita','association of research library','pc based','interlibrary loan system','library','e resource','patron facing','saas','platform','r.r.bowker','catalog','follett software','follett software','self service','cambridge information group','fretwell downing','library','innovative interface','uk based','product','product','service','sign on','virtual document eXchange','v insight','ebsco discovery service','skies','web scale','system','shem tov','service','service oriented','service','search','self check','self checkout','self hosted','serve','relevancy based','relevancy ranked','open educational resource','makerspace','north carolina state university','library academic','library bibliocommons','library cardcom','library','library companion','library company','library ex','library oriented','library public','library special','location based service','information system','library in','well established','full featured','industry','inn*reach','taos','web native','cloud based','library','central discovery index','migration','non us','service oriented','family owned','first search','first time','geac','functional requirements for bibliographic records','federal trade commission','gis','library provide','library related','longer term','kansas','niso circulation interchange protocol','new york public library','opac','year end','amsterdam  based','vtls','e resource','digital first','ever increasing','librarian','ils','integrated','library decision','library','library on','peer to peer','sector','subscription based','service oriented architecture','top tier','unified resource management','all web based','archive library museum','alma','book it','broad based','company in','development in','dos based','db/textworks','dra','ever growing','headquarter','impact online','impact verso','eos international','host terminal','migration'
]
except_word=['carl.x','eos.web','u.s','r.r.bowker','z39.50','non-u.s','r.r','carl.solution','absys.net']
tk=MWETokenizer()

shaicha2_path='D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\筛查记录2.6_2022.04.18.xlsx'
def readexcel(path,sheetnum):
    readbook = openpyxl.load_workbook(path)
    sheetname=readbook.get_sheet_names()
    sheet=readbook.get_sheet_by_name(sheetname[sheetnum])
    return sheet
ws=readexcel(shaicha2_path,0)
replace2=[]
replace2_word=[]
for w in range(2,ws.max_row+1):
    prase=ws.cell(w,1).value
    if prase is not None:
        replace2.append(prase)
    prase2=ws.cell(w,2).value
    if prase is not None:
        replace2_word.append(prase2)

ws3=readexcel(shaicha2_path,4)
replace3=[]
replace3_word=[]
for w in range(2,ws3.max_row+1):
    prase=ws3.cell(w,1).value
    if prase is not None:
        replace3.append(prase)
    prase2=ws3.cell(w,2).value
    if prase2 is not None:
        replace3_word.append(prase2)
print(replace3_word)

def findword(tt_tokens,word,artnum):
    for t_ in range(len(tt_tokens)):
        if tt_tokens[t_] == word:
            index = t_
            print('{} found in article {}:'.format(word, artnum + 1))
            s = index - 5 if index - 5 >= 0 else 0
            e = index + 5 if index + 5 <= len(tt_tokens) else len(tt_tokens)
            print('———— {}'.format(' '.join(tt_tokens[s:e])))

invalid_word=[]
invalid_sheet=readexcel(shaicha2_path,3)
for i in range(2,invalid_sheet.max_row+1):
    invalid_word.append(invalid_sheet.cell(i,1).value)


for ci in cizu:
    tk.add_mwe(ci)
punctuation_map2=[char for char in string.punctuation]
for i in range(testnum[1]-1,testnum[2]):
    path='D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\科研训练txt\\2.industry report全文txt\\'
    filepath=path+('%d' % (i+1))+'.txt'
    with open(filepath,'r',encoding='utf-8')as f:
        str1=f.read()
        str1=str1.split('\n')
        for s in str1:
            if s!='':
                str2 = re.sub('[\x00-\x1F\x7F-\x9F]', '', s)
                # str2=re.sub('\.',' ',str2)
                # without_punctuation=str2.translate(punctuation_map)
                lower_text = str2.lower()
                tokens = nltk.word_tokenize(lower_text)
                tag = nltk.pos_tag(tokens)
                wordlist = []
                for t in tag:
                    pos = get_wordnet_pos(t[1])
                    if pos != None:
                        new_word = lmtzr.lemmatize(t[0], pos)
                        wordlist.append(new_word)

                ttt = []
                for m in wordlist:
                    if m not in stopw:
                        if m not in punctuation_map2:
                            if m not in except_word:
                                if re.search('\.', m):
                                    m2 = re.sub('\.', ' ', m)
                                    m2 = m2.split(' ')
                                    m3 = nltk.pos_tag(m2)
                                    for mm in m3:
                                        pos = get_wordnet_pos(mm[1])
                                        if pos != None:
                                            new_word = lmtzr.lemmatize(mm[0], pos)
                                            if new_word in replace:
                                                ii = replace.index(new_word)
                                                mm2 = replace_word[ii]
                                                ttt.append(mm2)
                                            else:
                                                ttt.append(new_word)
                                else:
                                    if m in replace:
                                        ii = replace.index(m)
                                        m1 = replace_word[ii]
                                        ttt.append(m1)
                                    else:
                                        ttt.append(m)
                            else:
                                ttt.append(m)
                tt = ' '.join(ttt)
                tt_tokens = tk.tokenize(tt.split())
                # word_to_find='vtls'
                # findword(tt_tokens,word_to_find,i)  ###检索词
                new_tt_tokens = []
                for tt_ in tt_tokens:
                    newtt_=tt_
                    if newtt_ in replace2:
                        index = replace2.index(newtt_)
                        newtt_ = replace2_word[index]
                    if newtt_ in replace3:
                        index = replace3.index(newtt_)
                        newtt_ = replace3_word[index]
                    if newtt_ not in invalid_word:
                        new_tt_tokens.append(newtt_)
                wordnum.append(len(new_tt_tokens))
                tttt = ' '.join(new_tt_tokens)
                text.append(tttt)




max_f=None

vectorizer=CountVectorizer(min_df=1,token_pattern=r'\b\S+\b',max_features=max_f)
y=vectorizer.fit_transform(text)


y_vocabulary=vectorizer.vocabulary_
y_array=y.toarray()
y1=np.array(y_array[0])
for i in range(1,len(y_array)):
    y1+=y_array[i]

yy=[]
for yyy in sorted(y_vocabulary.items(),key=lambda x:x[0],reverse=False):
    yy.append(yyy[0])
print(len(y_array))

def wordcout(yy,y1,testnum,frequency=0):
    with open(f'D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\2022.08.14\\wordcount{testnum[0]}.csv','w',encoding='utf-8',newline='')as f:
        w=csv.writer(f)
        w.writerow(['Id','Label','weight'])
        for i in range(len(yy)):
            if y1[i] > frequency:
                    w.writerow([yy[i],yy[i],y1[i]])



from word_matrix import coocurance
def node_edge_BasePara(yy,y1,y_array,nodethreshnum,testnum,edgethreshnum):   #不借助于词频的
    para_word=[]
    for y_ in y_array:
        everypara_word=[]
        for y__ in range(len(y_)):
            if y_[y__]!=0:
                if y1[y__]>=nodethreshnum:
                    everypara_word.append(yy[y__])
        para_word.append(everypara_word)


        #print(u"-------这里输出第", j, u"类文本tf-idf权重前20的词语------")
        #print(L)

    node_str, edge_str = coocurance.build_matrix(para_word, True,edgethreshnum=edgethreshnum)
    with open(f'D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\2022.08.14\\edge{testnum}.csv', 'w', encoding='utf-8') as csv:
        csv.write("Source,Target,Weight\r")
        csv.write(edge_str)
    with open(f'D:\\YUAN JC\\#1#资讯管理学院\\2021-2022    大三图书馆学（下学期）\\科研学术训练\\2022.08.14\\node{testnum}.csv', 'w', encoding='utf-8') as csv:
        csv.write("Id,Label,weight\r")
        csv.write(node_str)




valid_word=[]
valid_sheet=readexcel(shaicha2_path,2)
for i in range(2,valid_sheet.max_row+1):
    valid_word.append(valid_sheet.cell(i,1).value)

wordcout(yy,y1,testnum=testnum,frequency=4)
node_edge_BasePara(yy,y1,y_array,nodethreshnum=5,testnum=testnum[0],edgethreshnum=4)










